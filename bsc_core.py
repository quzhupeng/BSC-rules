#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
平衡计分卡KPI数据处理核心类
功能：将非结构化的KPI指标数据转化为标准化的平衡计分卡格式
"""

import pandas as pd
import re
import openpyxl
from typing import Tuple, Dict, Optional, Union
from io import BytesIO


class BSCProcessor:
    """平衡计分卡数据处理器核心类"""

    def __init__(self, input_file=None):
        """
        初始化处理器

        Args:
            input_file: 输入Excel文件路径或BytesIO对象
        """
        self.input_file = input_file
        self.df = None
        self.target_col = None
        self.rule_col = None
        self.semi_target_col = None  # 半年度目标值列
        self.semi_rule_col = None    # 半年度计分规则列
        self._sheet_name = None  # 当前处理的sheet名（多sheet模式下使用）
        self.status_messages = []  # 用于存储处理状态信息

    def _log(self, message: str):
        """记录日志信息"""
        self.status_messages.append(message)

    def load_data(self) -> pd.DataFrame:
        """读取Excel数据"""
        try:
            self.df = pd.read_excel(self.input_file)
            self._log(f"成功读取文件，共 {len(self.df)} 行，{len(self.df.columns)} 列")
            self._log(f"列名: {list(self.df.columns)}")
            return self.df
        except Exception as e:
            raise Exception(f"读取文件失败: {e}")

    def identify_columns(self) -> Tuple[str, str]:
        """
        自动识别目标值列和计分规则列
        优先级：全年 > 年度 > 其他，排除所有半年相关列

        Returns:
            (目标值列名, 计分规则列名)
        """
        # 目标值关键字（按优先级排序：全年 > 年度 > 其他）
        target_keywords = [
            '全年目标值',
            '年度目标值', '2026目标值', '26年度目标值',
            '目标值', '考核目标值', 'kpi目标值', '指标值', '目标'
        ]
        # 计分规则关键字（按优先级排序：全年 > 年度 > 其他）
        rule_keywords = [
            '全年计分规则',
            '计分规则', '年度计分规则',
            '评分规则', '考核规则', '计分标准'
        ]

        # 需要排除的关键词（半年相关）
        exclude_keywords = ['半年度', '半年', '半期', '中期']

        def is_excluded(col_str: str) -> bool:
            """检查列名是否包含排除关键词"""
            return any(kw in col_str for kw in exclude_keywords)

        # 检查表头在哪一行（最多检查前3行）
        header_row_idx = None
        max_check_rows = min(4, len(self.df) + 1)

        # 首先检查当前列名（第0行，作为表头）
        for keyword in target_keywords:
            for col in self.df.columns:
                col_str = str(col)
                # 匹配关键字，但排除半年相关列
                if keyword in col_str and not is_excluded(col_str):
                    if header_row_idx is None:
                        header_row_idx = 0
                    break
            if header_row_idx == 0:
                break

        for keyword in rule_keywords:
            for col in self.df.columns:
                col_str = str(col)
                # 匹配关键字，但排除半年相关列
                if keyword in col_str and not is_excluded(col_str):
                    if header_row_idx is None:
                        header_row_idx = 0
                    break
            if header_row_idx == 0:
                break

        # 如果第一行列名没找到所有关键字，检查数据行
        if header_row_idx is None:
            found_target_in_col = any(
                any(kw in str(col) and not is_excluded(str(col)) for kw in target_keywords)
                for col in self.df.columns
            )
            found_rule_in_col = any(
                any(kw in str(col) and not is_excluded(str(col)) for kw in rule_keywords)
                for col in self.df.columns
            )

            for i in range(min(3, len(self.df))):
                row_data = self.df.iloc[i].astype(str)
                has_target = any(
                    any(kw in str(val) for kw in target_keywords)
                    for val in row_data.values
                )
                has_rule = any(
                    any(kw in str(val) for kw in rule_keywords)
                    for val in row_data.values
                )

                if has_target and has_rule:
                    header_row_idx = i + 1
                    break
                if has_target and not found_target_in_col:
                    header_row_idx = i + 1
                    found_target_in_col = True
                if has_rule and not found_rule_in_col:
                    if header_row_idx is None:
                        header_row_idx = i + 1
                    found_rule_in_col = True

        # 如果表头不在第0行，重新读取数据
        if header_row_idx and header_row_idx > 0:
            self._log(f"识别到表头在第{header_row_idx + 1}行")
            read_kwargs = {'header': header_row_idx}
            if self._sheet_name:
                read_kwargs['sheet_name'] = self._sheet_name
            self.df = pd.read_excel(self.input_file, **read_kwargs)
            self._log(f"列名: {list(self.df.columns)}")

        # 识别目标值列（按优先级）
        self.target_col = None
        for keyword in target_keywords:
            for col in self.df.columns:
                col_str = str(col)
                # 匹配关键字，排除半年相关列
                if keyword in col_str and not is_excluded(col_str):
                    self.target_col = col
                    self._log(f"找到目标值列: {self.target_col} (匹配关键字: {keyword})")
                    break
            if self.target_col:
                break

        # 识别计分规则列（按优先级）
        self.rule_col = None
        for keyword in rule_keywords:
            for col in self.df.columns:
                col_str = str(col)
                # 匹配关键字，排除半年相关列
                if keyword in col_str and not is_excluded(col_str):
                    self.rule_col = col
                    self._log(f"找到计分规则列: {self.rule_col} (匹配关键字: {keyword})")
                    break
            if self.rule_col:
                break

        if not self.target_col:
            raise Exception("无法识别目标值列，请检查前3行是否有包含'目标值'或'全年目标值'的列")
        if not self.rule_col:
            raise Exception("无法识别计分规则列，请检查前3行是否有包含'计分规则'或'全年计分规则'的列")

        self._log(f"最终使用 - 目标值列: {self.target_col}")
        self._log(f"最终使用 - 计分规则列: {self.rule_col}")

        return self.target_col, self.rule_col

    def identify_semi_annual_columns(self):
        """
        识别半年度目标值列和计分规则列（可选功能）
        找不到时不报错，设为None
        """
        semi_target_keywords = ['半年度目标值', '半年目标值', '中期目标值', '半期目标值']
        semi_rule_keywords = ['半年度计分规则', '半年计分规则', '中期计分规则', '半期计分规则']

        # 识别半年度目标值列
        self.semi_target_col = None
        for keyword in semi_target_keywords:
            for col in self.df.columns:
                if keyword in str(col):
                    self.semi_target_col = col
                    self._log(f"找到半年度目标值列: {self.semi_target_col} (匹配关键字: {keyword})")
                    break
            if self.semi_target_col:
                break

        # 识别半年度计分规则列
        self.semi_rule_col = None
        for keyword in semi_rule_keywords:
            for col in self.df.columns:
                if keyword in str(col):
                    self.semi_rule_col = col
                    self._log(f"找到半年度计分规则列: {self.semi_rule_col} (匹配关键字: {keyword})")
                    break
            if self.semi_rule_col:
                break

        if not self.semi_target_col:
            self._log("未找到半年度目标值列，跳过半年度处理")
        if not self.semi_rule_col:
            self._log("未找到半年度计分规则列，跳过半年度处理")

    @staticmethod
    def normalize_target_value(value: Union[str, float, int]) -> Tuple[float, bool]:
        """
        归一化目标值
        处理多种情况：
        1. 字符串 "85%" -> 0.85, is_percent=True
        2. 字符串 "85分" -> 85, is_percent=False
        3. 字符串 "10个" / "5人" / "3万" 等 -> 去除单位后转换为数字
        4. 数字 0.85 或 90 -> 对应浮点数, is_percent=False

        支持的单位：分、个、人、份、例、种、场、万、千、次、项、元、起、件、台、套、吨、株、亩、公斤、千克、立方米、平米、平方米、㎡、m²、m³

        Args:
            value: 原始值

        Returns:
            (标准化后的浮点数, 是否为百分比格式)
        """
        if pd.isna(value):
            return 0.0, False

        # 如果是字符串
        if isinstance(value, str):
            value = value.strip()
            # 去除逗号（如 1,000）
            value = value.replace(',', '')

            # 检查是否包含百分号
            if '%' in value:
                num_str = value.replace('%', '').strip()
                try:
                    return float(num_str) / 100, True
                except ValueError:
                    return 0.0, False

            # 需要去除的单位列表（百分号已在前面处理）
            units_to_remove = [
                '万元', '千元', '百元', '亿元',
                '分', '个', '人', '份', '例', '种', '场',
                '万', '千', '次', '项', '元', '起', '件',
                '台', '套', '吨', '株', '亩', '公斤', '千克',
                '立方米', '平米', '平方米', '㎡', 'm²', 'm³',
                '小时', '天', '日', '周', '月', '年',
                '公里', '千米', '米', 'm', 'km',
                '升', 'ml', 'l', 'g', 'kg',
                '分钟', '秒',
            ]

            # 去除所有单位（优先匹配长单位，避免部分匹配）
            for unit in sorted(units_to_remove, key=len, reverse=True):
                if value.endswith(unit):
                    value = value[:-len(unit)].strip()
                    break

            # 尝试直接转换
            try:
                return float(value), False
            except ValueError:
                return 0.0, False

        # 如果是数字
        elif isinstance(value, (int, float)):
            return float(value), False

        return 0.0, False

    @staticmethod
    def extract_explicit_baseline(rule_text: str) -> Optional[Tuple[float, str]]:
        """
        逻辑A：从规则文本中提取显式的底线值
        """
        if pd.isna(rule_text) or not isinstance(rule_text, str):
            return None

        rule_text = rule_text.strip()

        # 先检查是否包含得分阈值描述
        score_threshold_pattern = r'(?:低于|不足|少于)\s*([0-9]+)\s*分\s*(?:不得分|为0|得0分)'
        if re.search(score_threshold_pattern, rule_text):
            ratio_keywords = [
                r'实际\s*[:：]?\s*目标',
                r'达成\s*/\s*目标',
                r'除\s*以\s*目标',
                r'/\s*目标',
                r'÷\s*目标',
                r'最多\s*100分',
            ]
            if any(re.search(kw, rule_text) for kw in ratio_keywords):
                return None

        # 模式1: 低于/小于XX%不得分/得0分 (正向指标)
        patterns_positive = [
            r'(?:低于|小于)\s*([0-9]+\.?[0-9]*)%\s*(?:不得分|得0分)',
            r'<\s*([0-9]+\.?[0-9]*)%\s*(?:不得分|得0分)',
            r'(?:低于|小于)\s*([0-9]+\.?[0-9]*)%\s*分?\s*(?:不得分|得0分)',
            r'<\s*([0-9]+\.?[0-9]*)%\s*分?\s*(?:不得分|得0分)',
        ]

        for pattern in patterns_positive:
            match = re.search(pattern, rule_text)
            if match:
                value = float(match.group(1))
                if '%' in match.group(0):
                    return value / 100, '正向'
                return value, '正向'

        # 模式2: 高于/大于/超过XX%不得分/得0分 (逆向指标)
        patterns_negative = [
            r'(?:高于|大于|超过)\s*([0-9]+\.?[0-9]*)%\s*(?:不得分|得0分)',
            r'>\s*([0-9]+\.?[0-9]*)%\s*(?:不得分|得0分)',
            r'(?:高于|大于|超过)\s*([0-9]+\.?[0-9]*)%\s*分?\s*(?:不得分|得0分)',
            r'>\s*([0-9]+\.?[0-9]*)%\s*分?\s*(?:不得分|得0分)',
        ]

        for pattern in patterns_negative:
            match = re.search(pattern, rule_text)
            if match:
                value = float(match.group(1))
                if '%' in match.group(0):
                    return value / 100, '逆向'
                return value, '逆向'

        # 模式3: XX得60分/为60分（多级计分规则的底线值提取）
        # 格式如：=2400万，得60分；=4%，得60分；=12，得60分；=6，得60分
        # 使用非贪婪匹配，允许中间有其他字符
        patterns_60 = [
            # 带等号和单位的格式：=XX万/个/次/项/元/%，...得60分
            r'=([0-9]+\.?[0-9]*)\s*(?:万|个|次|项|元|%)?\s*[^0-9.]*?得60分',
            # 带等号和单位的格式，用逗号分隔：=XX万/个/次/项/元/%，...得60分
            r'=([0-9]+\.?[0-9]*)\s*(?:万|个|次|项|元|%)?，.*?得60分',
            # 不带等号，有单位的格式：XX万/个/次/项/元/%，...得60分
            r'(?:^|[^0-9.])([0-9]+\.?[0-9]*)\s*(?:万|个|次|项|元|%) [^0-9.]*?得60分',
            # 纯数字格式：XX，...得60分（排除百分比格式）
            r'(?:^|[^0-9.])([0-9]+)\s*，.*?得60分',
            # 通用格式：XX得60分（放在最后，作为兜底）
            r'([0-9]+\.?[0-9]*)\s*得60分',
        ]

        for pattern in patterns_60:
            match = re.search(pattern, rule_text)
            if match:
                value = float(match.group(1))
                if '%' in match.group(0):
                    return value / 100, '正向'
                return value, '正向'

        return None

    @staticmethod
    def extract_ratio_baseline(rule_text: str) -> Optional[Tuple[float, str]]:
        """提取比例型计分规则的得分阈值"""
        if pd.isna(rule_text) or not isinstance(rule_text, str):
            return None

        rule_text = rule_text.strip()

        # 检查是否是比例型规则
        ratio_keywords = [
            r'实际\s*[:：]?\s*目标',
            r'达成\s*/\s*目标',
            r'除\s*以\s*目标',
            r'/\s*目标',
            r'÷\s*目标',
        ]

        is_ratio_rule = any(re.search(kw, rule_text) for kw in ratio_keywords)
        is_max_score = re.search(r'最多\s*100分', rule_text) is not None

        if not is_ratio_rule and not is_max_score:
            return None

        # 提取得分阈值
        score_threshold_patterns = [
            r'低于\s*([0-9]+)\s*分[,，]?\s*(?:不得分|为0|得0分)',
            r'(?:不足|少于)\s*([0-9]+)\s*分[,，]?\s*(?:不得分|为0|得0分)',
            r'([0-9]+)分\s*(?:以下|为0)\s*(?:不得分|得0分)',
            r'满\s*100分.*?(?:低于|不足)\s*([0-9]+)\s*分[,，]?\s*不得分',
        ]

        for pattern in score_threshold_patterns:
            match = re.search(pattern, rule_text)
            if match:
                threshold = float(match.group(1))
                ratio = threshold / 100
                return (ratio, '正向')

        return (0.6, '正向')

    @staticmethod
    def extract_deduction_params(rule_text: str) -> Optional[Tuple[float, float, str, bool]]:
        """从规则文本中提取扣分参数"""
        if pd.isna(rule_text) or not isinstance(rule_text, str):
            return None

        rule_text = rule_text.strip()

        # 正向指标：每低/每差/每降/每少/每低于目标值/每起
        # 支持"扣"和"减"两种表达方式
        # 格式如：每低1%扣2分、每低于1%扣2分、每低于目标值2%，减5分、每低3分减7分、每少0.1%扣5分、每起扣10分
        # 注意：alternation要按长度降序排列（长的在前）避免部分匹配问题
        patterns_positive = [
            # 格式1: 每低于目标值X%[,，]扣/减Y分
            r'每低于目标值\s*([0-9]+\.?[0-9]*)%\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式2: 每低于X%[,，]扣/减Y分
            r'每低于\s*([0-9]+\.?[0-9]*)%\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式3: 每低X%[,，]扣/减Y分
            r'每低\s*([0-9]+\.?[0-9]*)%\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式4: 每少X%[,，]扣/减Y分（如：每少0.1%扣5分）
            r'每少\s*([0-9]+\.?[0-9]*)%\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式5: 每低于目标值X分[,，]扣/减Y分
            r'每低于目标值\s*([0-9]+\.?[0-9]*)\s*分\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式6: 每低X分[,，]扣/减Y分
            r'每低\s*([0-9]+\.?[0-9]*)\s*分\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式7: 每少X个[,，]扣/减Y分
            r'每少\s*([0-9]+\.?[0-9]*)\s*个\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式8: 每起X扣/减Y分 或 每起扣/减Y分（逆向指标：如"每起扣10分"表示事故类指标）
            # 这个格式需要特殊处理，因为默认是逆向指标（越少越好）
            r'每起\s*(?:([0-9]+\.?[0-9]*)\s*(?:个|单位|起)?\s*)?(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式9: 每X起扣/减Y分（如：每1起扣10分，逆向指标）
            r'每\s*([0-9]+\.?[0-9]*)\s*起\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            # 格式9: 每差/每降/每小 X单位[,，]扣/减Y分
            r'每[差小降](?:于目标值)?\s*([0-9]+\.?[0-9]*)\s*(?:%|[个人次项元万千百])?\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
        ]

        for pattern in patterns_positive:
            match = re.search(pattern, rule_text)
            if match:
                x_str = match.group(1)
                y_str = match.group(2)
                # 处理x为空的情况（如"每起扣10分"，默认每起算1个单位）
                if not x_str or x_str == '':
                    x = 1.0
                    y = float(y_str)
                else:
                    x = float(x_str)
                    y = float(y_str)
                # 检查匹配的文本中是否包含%
                matched_text = match.group(0)
                has_percent = '%' in matched_text
                return (x, y, '正向', has_percent)

        # 逆向指标：每高/每超/每多/每高于目标值
        patterns_negative = [
            r'每高于目标值\s*([0-9]+\.?[0-9]*)%\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            r'每高\s*([0-9]+\.?[0-9]*)%\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            r'每高于目标值\s*([0-9]+\.?[0-9]*)\s*分\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            r'每高\s*([0-9]+\.?[0-9]*)\s*分\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            r'每[高超多](?:于目标值)?\s*([0-9]+\.?[0-9]*)\s*(?:%|[个人次项元万千百])?\s*[,，]?\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
        ]

        for pattern in patterns_negative:
            match = re.search(pattern, rule_text)
            if match:
                x = float(match.group(1))
                y = float(match.group(2))
                matched_text = match.group(0)
                has_percent = '%' in matched_text
                return (x, y, '逆向', has_percent)

        return None

    @staticmethod
    def extract_accident_params(rule_text: str) -> Optional[Tuple[float, float, str]]:
        """
        提取"每起扣X分"类型的参数（逆向指标：事故类）

        这类指标的逻辑是：从0开始（满分），每有1起事故扣X分
        底线值（得60分）= 允许扣40分时的事故数 = 40 / X

        Args:
            rule_text: 规则文本

        Returns:
            (每起扣X分, 方向) 或 None
        """
        if pd.isna(rule_text) or not isinstance(rule_text, str):
            return None

        rule_text = rule_text.strip()

        # 匹配"每起扣X分"或"每X起扣Y分"
        patterns = [
            r'每起\s*(?:([0-9]+\.?[0-9]*)\s*(?:个|单位|起)?\s*)?(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
            r'每\s*([0-9]+\.?[0-9]*)\s*起\s*(?:扣|减)\s*([0-9]+\.?[0-9]*)\s*分',
        ]

        for pattern in patterns:
            match = re.search(pattern, rule_text)
            if match:
                x_str = match.group(1)  # 每X起（可能为空）
                y_str = match.group(2) if len(match.groups()) >= 2 else match.group(1)  # 扣Y分
                # 如果x_str为空，说明是"每起扣Y分"，默认每1起扣Y分
                if not x_str or x_str == '':
                    y = float(y_str)
                    x = 1.0
                else:
                    x = float(x_str)
                    y = float(y_str)
                return (x, y, '逆向')

        return None

    @staticmethod
    def detect_indicator_direction(rule_text: str) -> Optional[str]:
        """从规则文本中检测指标方向（正向/逆向）"""
        if pd.isna(rule_text) or not isinstance(rule_text, str):
            return None

        # 逆向指标关键词
        negative_keywords = [
            '投诉率', '差错率', '故障率', '缺陷率', '不良率', '报废率',
            '成本控制', '控制在', '不超过', '不高于', '每高', '每超',
            '下降', '降低', '越低', '超出', '超支',
        ]

        # 正向指标关键词
        positive_keywords = [
            '完成率', '达成率', '实现率', '增长率', '提升率',
            '收入', '利润', '销售额', '产量', '达标', '超额',
            '越高', '不少于', '每低', '每降', '每差', '每少',
        ]

        positive_score = sum(1 for kw in positive_keywords if kw in rule_text)
        negative_score = sum(1 for kw in negative_keywords if kw in rule_text)

        if positive_score > negative_score:
            return '正向'
        elif negative_score > positive_score:
            return '逆向'
        return None

    def calculate_baseline(self, target: float, rule_text: str,
                          is_percent: bool) -> Tuple[float, str, str]:
        """计算底线值（核心函数）"""
        baseline = None
        status = '成功'
        direction = self.detect_indicator_direction(rule_text) or '正向'

        # 逻辑A（最高优先级）: 事故类指标（每起扣X分，逆向）
        # 这类指标从0开始（满分），每有1起事故扣X分
        # 底线值（得60分）= 允许扣40分时的事故数 = 40 / X
        accident_params = self.extract_accident_params(rule_text)
        if accident_params is not None:
            x, y, accident_params_direction = accident_params
            direction = accident_params_direction
            # 从0开始，每X起扣Y分，得60分时可以有多少起？
            # 40分 / Y分 = 允许扣的次数
            # 每次扣X起，所以允许起数 = (40 / Y) × X
            allowed_count = (40 / y) * x
            baseline = allowed_count
            return baseline, '成功', direction

        # 逻辑B: 比例型规则
        ratio_info = self.extract_ratio_baseline(rule_text)
        if ratio_info is not None:
            ratio, detected_direction = ratio_info
            direction = detected_direction
            baseline = target * ratio
            return baseline, '成功', direction

        # 逻辑C: 扣分推导
        deduction = self.extract_deduction_params(rule_text)
        if deduction is not None:
            x, y, deduction_direction, rule_has_percent = deduction
            direction = deduction_direction

            allowed_gap = (40 / y) * x

            if rule_has_percent:
                allowed_gap = allowed_gap / 100

            if direction == '逆向':
                baseline = target + allowed_gap
            else:
                baseline = target - allowed_gap

            return baseline, '成功', direction

        # 逻辑C: 显式阈值
        explicit_baseline = self.extract_explicit_baseline(rule_text)
        if explicit_baseline is not None:
            baseline, detected_direction = explicit_baseline
            if not (
                abs(baseline - 60) < 0.01 and '不得分' in rule_text
            ):
                if detected_direction:
                    direction = detected_direction
                return baseline, '成功', direction

            return baseline, '成功', direction

        # 逻辑D: 默认兜底
        if direction == '逆向':
            baseline = target * 1.2
        else:
            baseline = target * 0.8

        status = '人工校验'
        return baseline, status, direction

    @staticmethod
    def format_value(value: float, is_percent: bool) -> str:
        """将数值格式化为显示字符串"""
        if is_percent:
            return f"{value * 100:.2f}%"
        else:
            if value == int(value):
                return str(int(value))
            elif abs(value) >= 1000:
                return f"{value:.2f}"
            elif abs(value) >= 100:
                return f"{value:.2f}"
            else:
                return f"{value:.4f}".rstrip('0').rstrip('.')

    def generate_standard_rule(self, target: float, baseline: float,
                               direction: str, is_percent: bool) -> str:
        """生成规范化计分规则文案"""
        target_str = self.format_value(target, is_percent)
        baseline_str = self.format_value(baseline, is_percent)

        if direction == '逆向':
            template = (
                f"P为指标实际值，{target_str}为目标值，{baseline_str}为底线值。\n"
                f"1.若P≤{target_str}，得100分（满分）；\n"
                f"2.若{target_str}<P<{baseline_str}，按线性比例计算，即："
                f"得分=100-(P-{target_str})/({baseline_str}-{target_str})×(100-60)；\n"
                f"3.若P={baseline_str}，得60分（基础分）；\n"
                f"4.若P＞{baseline_str}，得0分。"
            )
        else:
            template = (
                f"P为指标实际值，{target_str}为目标值，{baseline_str}为底线值。\n"
                f"1.若P≥{target_str}，得100分（满分）；\n"
                f"2.若{baseline_str}<P<{target_str}，按线性比例计算，即："
                f"得分=60+(P-{baseline_str})/({target_str}-{baseline_str})×(100-60)；\n"
                f"3.若P={baseline_str}，得60分（基础分）；\n"
                f"4.若P<{baseline_str}，得0分。"
            )

        return template

    def process_row(self, row: pd.Series, target_col=None, rule_col=None) -> Dict:
        """处理单行数据"""
        tc = target_col if target_col is not None else self.target_col
        rc = rule_col if rule_col is not None else self.rule_col
        target_raw = row[tc]
        rule_text = row[rc] if rc in row else ""

        target, is_percent = self.normalize_target_value(target_raw)
        baseline, status, direction = self.calculate_baseline(
            target, str(rule_text), is_percent
        )

        if status == '人工校验':
            if baseline < target:
                direction = '正向'
            elif baseline > target:
                direction = '逆向'

        standard_rule = self.generate_standard_rule(
            target, baseline, direction, is_percent
        )

        baseline_display = self.format_value(baseline, is_percent)

        return {
            '推导底线值': baseline_display,
            '规范版计分规则': standard_rule,
            '解析状态': status,
            '指标方向': direction,
            '目标值_归一化': target,
            '底线值_归一化': baseline,
            '是否百分比': is_percent
        }

    def process(self, progress_callback=None) -> pd.DataFrame:
        """
        执行完整处理流程

        Args:
            progress_callback: 进度回调函数，参数为当前进度(0-100)

        Returns:
            处理后的DataFrame
        """
        self.status_messages.clear()

        # 加载数据
        self.load_data()
        if progress_callback:
            progress_callback(20)

        # 识别列
        self.identify_columns()
        self.identify_semi_annual_columns()
        if progress_callback:
            progress_callback(40)

        # 添加新列
        results = {
            '推导底线值': [],
            '规范版计分规则': [],
            '解析状态': [],
            '指标方向': []
        }

        total_rows = len(self.df)
        for idx, row in self.df.iterrows():
            try:
                result = self.process_row(row)
                results['推导底线值'].append(result['推导底线值'])
                results['规范版计分规则'].append(result['规范版计分规则'])
                results['解析状态'].append(result['解析状态'])
                results['指标方向'].append(result['指标方向'])
            except Exception as e:
                self._log(f"警告: 第{idx+2}行处理失败: {e}")
                results['推导底线值'].append('ERROR')
                results['规范版计分规则'].append(f'解析失败: {str(e)}')
                results['解析状态'].append(f'ERROR: {str(e)[:50]}')
                results['指标方向'].append('unknown')

            # 更新进度
            if progress_callback:
                progress = 40 + int((idx + 1) / total_rows * 50)
                progress_callback(progress)

        # 将结果添加到原DataFrame
        for col, values in results.items():
            self.df[col] = values

        # 半年度处理
        if self.semi_target_col and self.semi_rule_col:
            self._log("开始半年度数据处理...")
            semi_results = {
                '半年度_推导底线值': [],
                '半年度_规范版计分规则': [],
                '半年度_解析状态': [],
                '半年度_指标方向': []
            }

            for idx, row in self.df.iterrows():
                # 检查半年度目标值是否为空
                semi_target_raw = row[self.semi_target_col]
                if pd.isna(semi_target_raw) or str(semi_target_raw).strip() == '':
                    semi_results['半年度_推导底线值'].append('')
                    semi_results['半年度_规范版计分规则'].append('')
                    semi_results['半年度_解析状态'].append('无半年度数据')
                    semi_results['半年度_指标方向'].append('')
                    continue

                try:
                    result = self.process_row(row, target_col=self.semi_target_col, rule_col=self.semi_rule_col)
                    semi_results['半年度_推导底线值'].append(result['推导底线值'])
                    semi_results['半年度_规范版计分规则'].append(result['规范版计分规则'])
                    semi_results['半年度_解析状态'].append(result['解析状态'])
                    semi_results['半年度_指标方向'].append(result['指标方向'])
                except Exception as e:
                    self._log(f"警告: 第{idx+2}行半年度处理失败: {e}")
                    semi_results['半年度_推导底线值'].append('ERROR')
                    semi_results['半年度_规范版计分规则'].append(f'解析失败: {str(e)}')
                    semi_results['半年度_解析状态'].append(f'ERROR: {str(e)[:50]}')
                    semi_results['半年度_指标方向'].append('unknown')

            for col, values in semi_results.items():
                self.df[col] = values

            # 半年度统计
            semi_status_counts = self.df['半年度_解析状态'].value_counts()
            self._log("半年度处理完成！统计信息：")
            for status, count in semi_status_counts.items():
                self._log(f"  {status}: {count} 行")

        # 统计
        status_counts = self.df['解析状态'].value_counts()
        self._log("处理完成！统计信息：")
        for status, count in status_counts.items():
            self._log(f"  {status}: {count} 行")

        manual_check = (self.df['解析状态'] == '人工校验').sum()
        self._log(f"需要人工核对的行数: {manual_check}")

        if progress_callback:
            progress_callback(100)

        return self.df

    def save_to_bytesio(self) -> BytesIO:
        """
        将处理结果保存到BytesIO对象（用于Streamlit下载）

        Returns:
            BytesIO对象
        """
        if self.df is None:
            raise Exception("请先执行process()方法")

        output = BytesIO()

        # 设置列宽
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False)

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # 设置列宽
            column_widths = {
                '推导底线值': 15,
                '规范版计分规则': 60,
                '解析状态': 20,
                '指标方向': 12,
                '半年度_推导底线值': 15,
                '半年度_规范版计分规则': 60,
                '半年度_解析状态': 20,
                '半年度_指标方向': 12
            }

            for col, width in column_widths.items():
                if col in self.df.columns:
                    col_idx = list(self.df.columns).index(col) + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = width

            # 设置规范版计分规则列为自动换行
            for wrap_col in ['规范版计分规则', '半年度_规范版计分规则']:
                if wrap_col in self.df.columns:
                    col_idx = list(self.df.columns).index(wrap_col) + 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in worksheet[col_letter]:
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        output.seek(0)
        return output

    def get_stats(self) -> Dict:
        """获取处理统计信息"""
        if self.df is None or '解析状态' not in self.df.columns:
            return {}

        status_counts = self.df['解析状态'].value_counts()
        stats = {
            'total': len(self.df),
            'success': int(status_counts.get('成功', 0)),
            'manual_check': int(status_counts.get('人工校验', 0)),
            'error': int(sum(cnt for status, cnt in status_counts.items() if 'ERROR' in status)),
        }

        # 半年度统计
        if '半年度_解析状态' in self.df.columns:
            semi_counts = self.df['半年度_解析状态'].value_counts()
            stats['semi_annual'] = {
                'success': int(semi_counts.get('成功', 0)),
                'manual_check': int(semi_counts.get('人工校验', 0)),
                'no_data': int(semi_counts.get('无半年度数据', 0)),
                'error': int(sum(cnt for status, cnt in semi_counts.items() if 'ERROR' in status)),
            }

        return stats

    def get_logs(self) -> list:
        """获取处理日志"""
        return self.status_messages


class BSCMultiSheetProcessor:
    """
    多Sheet平衡计分卡数据处理器
    用于处理包含多个sheet的Excel文件，每个sheet可能包含KPI数据
    """

    def __init__(self, input_file):
        """
        初始化多Sheet处理器

        Args:
            input_file: 输入Excel文件路径或BytesIO对象
        """
        self.input_file = input_file
        self.results = {}  # {sheet_name: processed_df}
        self.stats = {}  # {sheet_name: stats}
        self.all_logs = []  # 所有日志
        self.success_sheets = []  # 成功处理的sheet名称
        self.failed_sheets = []  # 失败的sheet名称
        self.skipped_sheets = []  # 跳过的sheet名称（无有效列）

    def _log(self, message: str):
        """记录日志信息"""
        self.all_logs.append(message)

    def get_sheet_names(self) -> list:
        """获取Excel文件中所有sheet名称"""
        try:
            xl_file = pd.ExcelFile(self.input_file)
            return xl_file.sheet_names
        except Exception as e:
            self._log(f"读取sheet列表失败: {e}")
            return []

    def _check_sheet_has_valid_columns(self, sheet_name: str) -> bool:
        """
        检查sheet是否包含有效的目标值列和计分规则列

        Args:
            sheet_name: sheet名称

        Returns:
            是否包含有效列
        """
        try:
            df = pd.read_excel(self.input_file, sheet_name=sheet_name, header=None)
            df_str = df.astype(str)

            # 检查是否包含目标值关键字
            target_keywords = ['目标值', '年度目标值', '全年目标值', '2026目标值']
            has_target = False
            for kw in target_keywords:
                if df_str.apply(lambda row: row.str.contains(kw, na=False).any(), axis=1).any():
                    has_target = True
                    break

            # 检查是否包含计分规则关键字
            rule_keywords = ['计分规则', '评分规则', '考核规则', '计分标准']
            has_rule = False
            for kw in rule_keywords:
                if df_str.apply(lambda row: row.str.contains(kw, na=False).any(), axis=1).any():
                    has_rule = True
                    break

            return has_target and has_rule
        except Exception:
            return False

    def process(self, progress_callback=None) -> Dict:
        """
        处理所有sheet

        Args:
            progress_callback: 进度回调函数，参数为当前进度(0-100)

        Returns:
            处理结果统计字典
        """
        sheet_names = self.get_sheet_names()
        if not sheet_names:
            raise Exception("无法读取Excel文件的sheet列表")

        total_sheets = len(sheet_names)
        self._log(f"发现 {total_sheets} 个sheet: {', '.join(sheet_names)}")

        for idx, sheet_name in enumerate(sheet_names):
            if progress_callback:
                progress = int((idx / total_sheets) * 100)
                progress_callback(progress)

            self._log(f"\n正在处理 sheet: {sheet_name}")

            # 检查sheet是否包含有效列
            if not self._check_sheet_has_valid_columns(sheet_name):
                self._log(f"  ⚠️ Sheet '{sheet_name}' 不包含目标值或计分规则列，跳过")
                self.skipped_sheets.append(sheet_name)
                continue

            try:
                # 创建单sheet处理器
                processor = BSCProcessor(self.input_file)
                processor._sheet_name = sheet_name

                # 读取指定sheet的数据
                processor.df = pd.read_excel(self.input_file, sheet_name=sheet_name)

                # 识别列
                try:
                    processor.identify_columns()
                    processor.identify_semi_annual_columns()
                except Exception as e:
                    self._log(f"  ❌ Sheet '{sheet_name}' 列识别失败: {e}")
                    self.failed_sheets.append(sheet_name)
                    continue

                # 处理数据
                try:
                    result_df = processor._process_df(
                        processor.df, processor.target_col, processor.rule_col,
                        semi_target_col=processor.semi_target_col,
                        semi_rule_col=processor.semi_rule_col
                    )
                    self.results[sheet_name] = result_df
                    self.stats[sheet_name] = processor.get_stats()
                    self.success_sheets.append(sheet_name)

                    # 添加处理日志
                    for log in processor.get_logs():
                        self._log(f"  {log}")

                    status_counts = result_df['解析状态'].value_counts()
                    success_count = int(status_counts.get('成功', 0))
                    manual_count = int(status_counts.get('人工校验', 0))
                    error_count = int(sum(cnt for status, cnt in status_counts.items() if 'ERROR' in status))

                    self._log(f"  ✅ Sheet '{sheet_name}' 处理成功: {len(result_df)}行 "
                             f"(成功:{success_count}, 人工校验:{manual_count}, 错误:{error_count})")

                except Exception as e:
                    self._log(f"  ❌ Sheet '{sheet_name}' 数据处理失败: {e}")
                    self.failed_sheets.append(sheet_name)

            except Exception as e:
                self._log(f"  ❌ Sheet '{sheet_name}' 处理异常: {e}")
                self.failed_sheets.append(sheet_name)

        if progress_callback:
            progress_callback(100)

        # 生成汇总日志
        self._log("\n" + "=" * 70)
        self._log("多Sheet处理汇总:")
        self._log(f"  总sheet数: {total_sheets}")
        self._log(f"  ✅ 成功处理: {len(self.success_sheets)} 个")
        if self.success_sheets:
            self._log(f"     {', '.join(self.success_sheets)}")
        self._log(f"  ⚠️ 跳过（无有效列）: {len(self.skipped_sheets)} 个")
        if self.skipped_sheets:
            self._log(f"     {', '.join(self.skipped_sheets)}")
        self._log(f"  ❌ 处理失败: {len(self.failed_sheets)} 个")
        if self.failed_sheets:
            self._log(f"     {', '.join(self.failed_sheets)}")
        self._log("=" * 70)

        return {
            'total': total_sheets,
            'success': len(self.success_sheets),
            'skipped': len(self.skipped_sheets),
            'failed': len(self.failed_sheets),
            'sheet_names': sheet_names,
            'success_sheets': self.success_sheets,
            'skipped_sheets': self.skipped_sheets,
            'failed_sheets': self.failed_sheets,
        }

    def save_to_bytesio(self) -> BytesIO:
        """
        将所有成功处理的sheet保存到一个Excel文件

        Returns:
            BytesIO对象
        """
        if not self.results:
            raise Exception("没有成功处理的sheet数据")

        output = BytesIO()

        # 设置列宽
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in self.results.items():
                # Excel sheet名称不能超过31个字符，且不能包含特殊字符
                safe_sheet_name = sheet_name[:31]
                # 替换不允许的字符
                for char in ['\\', '/', '*', '[', ']', ':', '?']:
                    safe_sheet_name = safe_sheet_name.replace(char, '_')

                df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                # 设置列宽
                worksheet = writer.sheets[safe_sheet_name]
                column_widths = {
                    '推导底线值': 15,
                    '规范版计分规则': 60,
                    '解析状态': 20,
                    '指标方向': 12,
                    '半年度_推导底线值': 15,
                    '半年度_规范版计分规则': 60,
                    '半年度_解析状态': 20,
                    '半年度_指标方向': 12
                }

                for col, width in column_widths.items():
                    if col in df.columns:
                        col_idx = list(df.columns).index(col) + 1
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        worksheet.column_dimensions[col_letter].width = width

                # 设置规范版计分规则列为自动换行
                for wrap_col in ['规范版计分规则', '半年度_规范版计分规则']:
                    if wrap_col in df.columns:
                        col_idx = list(df.columns).index(wrap_col) + 1
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        for cell in worksheet[col_letter]:
                            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        output.seek(0)
        return output

    def get_stats(self) -> Dict:
        """
        获取所有sheet的处理统计

        Returns:
            统计字典
        """
        return {
            'summary': {
                'total': len(self.success_sheets) + len(self.failed_sheets) + len(self.skipped_sheets),
                'success': len(self.success_sheets),
                'skipped': len(self.skipped_sheets),
                'failed': len(self.failed_sheets),
            },
            'by_sheet': self.stats,
            'success_sheets': self.success_sheets,
            'failed_sheets': self.failed_sheets,
            'skipped_sheets': self.skipped_sheets,
        }

    def get_logs(self) -> list:
        """获取所有处理日志"""
        return self.all_logs


# 为BSCProcessor添加内部处理方法（用于多sheet处理）
def _process_df(self, df: pd.DataFrame, target_col: str, rule_col: str,
                semi_target_col=None, semi_rule_col=None) -> pd.DataFrame:
    """
    处理已加载数据框的内部方法
    """
    self.df = df
    self.target_col = target_col
    self.rule_col = rule_col

    # 添加新列
    results = {
        '推导底线值': [],
        '规范版计分规则': [],
        '解析状态': [],
        '指标方向': []
    }

    total_rows = len(self.df)
    for idx, row in self.df.iterrows():
        try:
            result = self.process_row(row)
            results['推导底线值'].append(result['推导底线值'])
            results['规范版计分规则'].append(result['规范版计分规则'])
            results['解析状态'].append(result['解析状态'])
            results['指标方向'].append(result['指标方向'])
        except Exception as e:
            self._log(f"警告: 第{idx+2}行处理失败: {e}")
            results['推导底线值'].append('ERROR')
            results['规范版计分规则'].append(f'解析失败: {str(e)}')
            results['解析状态'].append(f'ERROR: {str(e)[:50]}')
            results['指标方向'].append('unknown')

    # 将结果添加到原DataFrame
    for col, values in results.items():
        self.df[col] = values

    # 半年度处理
    if semi_target_col and semi_rule_col:
        self._log("开始半年度数据处理...")
        semi_results = {
            '半年度_推导底线值': [],
            '半年度_规范版计分规则': [],
            '半年度_解析状态': [],
            '半年度_指标方向': []
        }

        for idx, row in self.df.iterrows():
            semi_target_raw = row[semi_target_col]
            if pd.isna(semi_target_raw) or str(semi_target_raw).strip() == '':
                semi_results['半年度_推导底线值'].append('')
                semi_results['半年度_规范版计分规则'].append('')
                semi_results['半年度_解析状态'].append('无半年度数据')
                semi_results['半年度_指标方向'].append('')
                continue

            try:
                result = self.process_row(row, target_col=semi_target_col, rule_col=semi_rule_col)
                semi_results['半年度_推导底线值'].append(result['推导底线值'])
                semi_results['半年度_规范版计分规则'].append(result['规范版计分规则'])
                semi_results['半年度_解析状态'].append(result['解析状态'])
                semi_results['半年度_指标方向'].append(result['指标方向'])
            except Exception as e:
                self._log(f"警告: 第{idx+2}行半年度处理失败: {e}")
                semi_results['半年度_推导底线值'].append('ERROR')
                semi_results['半年度_规范版计分规则'].append(f'解析失败: {str(e)}')
                semi_results['半年度_解析状态'].append(f'ERROR: {str(e)[:50]}')
                semi_results['半年度_指标方向'].append('unknown')

        for col, values in semi_results.items():
            self.df[col] = values

        semi_status_counts = self.df['半年度_解析状态'].value_counts()
        self._log("半年度处理完成！统计信息：")
        for status, count in semi_status_counts.items():
            self._log(f"  {status}: {count} 行")

    # 统计
    status_counts = self.df['解析状态'].value_counts()
    self._log("处理完成！统计信息：")
    for status, count in status_counts.items():
        self._log(f"  {status}: {count} 行")

    manual_check = (self.df['解析状态'] == '人工校验').sum()
    self._log(f"需要人工核对的行数: {manual_check}")

    return self.df


# 动态添加方法到BSCProcessor类
BSCProcessor._process_df = _process_df


class BSCBatchProcessor:
    """批量处理多个Excel文件"""

    def __init__(self):
        self.file_results = {}      # {文件名: {sheet名: df}}
        self.file_stats = {}        # {文件名: summary_dict}
        self.all_logs = []
        self.success_files = []
        self.failed_files = []

    def _log(self, message: str):
        """记录日志信息"""
        self.all_logs.append(message)

    def process(self, files, progress_callback=None) -> dict:
        """
        批量处理多个Excel文件

        Args:
            files: list of (filename, BytesIO) 元组
            progress_callback: 进度回调函数，参数为当前进度(0-100)

        Returns:
            汇总dict: {total, success, failed, success_files, failed_files}
        """
        total_files = len(files)
        self._log(f"开始批量处理 {total_files} 个文件")

        for file_idx, (filename, file_bytes) in enumerate(files):
            if progress_callback:
                progress = int((file_idx / total_files) * 100)
                progress_callback(progress)

            self._log(f"\n{'='*70}")
            self._log(f"正在处理文件 [{file_idx+1}/{total_files}]: {filename}")
            self._log(f"{'='*70}")

            try:
                file_bytes.name = filename
                multi_processor = BSCMultiSheetProcessor(file_bytes)
                summary = multi_processor.process()

                # 收集日志
                for log in multi_processor.get_logs():
                    self._log(f"  {log}")

                if summary['success'] > 0:
                    self.file_results[filename] = multi_processor.results
                    self.file_stats[filename] = summary
                    self.success_files.append(filename)
                    self._log(f"✅ 文件 '{filename}' 处理成功: "
                             f"{summary['success']}个Sheet成功, "
                             f"{summary['skipped']}个跳过, "
                             f"{summary['failed']}个失败")
                else:
                    self.failed_files.append(filename)
                    self._log(f"❌ 文件 '{filename}' 无可处理的Sheet")

            except Exception as e:
                self.failed_files.append(filename)
                self._log(f"❌ 文件 '{filename}' 处理失败: {e}")

        if progress_callback:
            progress_callback(100)

        # 生成汇总日志
        self._log(f"\n{'='*70}")
        self._log("批量处理汇总:")
        self._log(f"  总文件数: {total_files}")
        self._log(f"  ✅ 成功: {len(self.success_files)} 个")
        if self.success_files:
            self._log(f"     {', '.join(self.success_files)}")
        self._log(f"  ❌ 失败: {len(self.failed_files)} 个")
        if self.failed_files:
            self._log(f"     {', '.join(self.failed_files)}")
        self._log(f"{'='*70}")

        return {
            'total': total_files,
            'success': len(self.success_files),
            'failed': len(self.failed_files),
            'success_files': self.success_files,
            'failed_files': self.failed_files,
        }

    def save_to_bytesio(self) -> BytesIO:
        """
        将所有成功处理的文件结果合并保存到一个Excel文件

        Sheet命名规则：{文件名去后缀}_{原Sheet名}
        - 截断到31字符（Excel限制）
        - 替换 \\ / * [ ] : ? 为 _
        - 冲突时加 _2, _3 后缀

        Returns:
            BytesIO对象
        """
        if not self.file_results:
            raise Exception("没有成功处理的文件数据")

        output = BytesIO()
        used_names = {}  # 用于检测Sheet名冲突

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for filename, sheets in self.file_results.items():
                # 去除文件后缀
                file_base = filename
                for ext in ['.xlsx', '.xls', '.XLSX', '.XLS']:
                    if file_base.endswith(ext):
                        file_base = file_base[:-len(ext)]
                        break

                for sheet_name, df in sheets.items():
                    # 构造Sheet名
                    raw_name = f"{file_base}_{sheet_name}"

                    # 替换不允许的字符
                    safe_name = raw_name
                    for char in ['\\', '/', '*', '[', ']', ':', '?']:
                        safe_name = safe_name.replace(char, '_')

                    # 截断到31字符
                    safe_name = safe_name[:31]

                    # 冲突处理
                    if safe_name in used_names:
                        used_names[safe_name] += 1
                        suffix = f"_{used_names[safe_name]}"
                        safe_name = safe_name[:31 - len(suffix)] + suffix
                    else:
                        used_names[safe_name] = 1

                    df.to_excel(writer, sheet_name=safe_name, index=False)

                    # 设置列宽和格式
                    worksheet = writer.sheets[safe_name]
                    column_widths = {
                        '推导底线值': 15,
                        '规范版计分规则': 60,
                        '解析状态': 20,
                        '指标方向': 12,
                        '半年度_推导底线值': 15,
                        '半年度_规范版计分规则': 60,
                        '半年度_解析状态': 20,
                        '半年度_指标方向': 12
                    }

                    for col, width in column_widths.items():
                        if col in df.columns:
                            col_idx = list(df.columns).index(col) + 1
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            worksheet.column_dimensions[col_letter].width = width

                    # 设置规范版计分规则列为自动换行
                    for wrap_col in ['规范版计分规则', '半年度_规范版计分规则']:
                        if wrap_col in df.columns:
                            col_idx = list(df.columns).index(wrap_col) + 1
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            for cell in worksheet[col_letter]:
                                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        output.seek(0)
        return output

    def get_logs(self) -> list:
        """获取所有处理日志"""
        return self.all_logs
